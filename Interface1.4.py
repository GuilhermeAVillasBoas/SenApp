
import os
import threading
import webbrowser
import pygetwindow as gw
import pyautogui
import subprocess
import time
import threading
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, simpledialog, filedialog
from openpyxl import load_workbook
import win32com.client
from datetime import datetime

# VARIÁVEIS
janela_sge = "TOTVS Linha RM - Serviços  Alias: CorporeRM | 3-SENAI"  
janela_mecLogin = "[MEC - SISTEC -v.4279 ] - Google Chrome"
janela_mec = "[MEC - SISTEC] - Google Chrome"
url_mec_login = "https://sistec.mec.gov.br/login/login"
chrome_path = webbrowser.get(using='windows-default')
popup = None
label_mensagem = None
botao_ok = None
nome_saida = None

#FUNÇÕES GERAIS

def salvar_com_nome_disponivel(nome_base="2.xlsx"):
    nome, ext = os.path.splitext(nome_base)
    contador = 1
    novo_nome = nome_base
    while os.path.exists(novo_nome):
        try:
            with open(novo_nome, 'a'):
                return novo_nome
        except PermissionError:
            novo_nome = f"{nome}_{contador}{ext}"
            contador += 1
    return novo_nome

def obter_intervalo_datas():
    def formatar_data(event, entrada):
        valor = entrada.get().replace("/", "")
        novo_valor = ""
        for i, c in enumerate(valor):
            if i == 2 or i == 4:
                novo_valor += "/"
            novo_valor += c
        entrada.delete(0, tk.END)
        entrada.insert(0, novo_valor[:10])

    def on_confirmar():
        try:
            data1 = datetime.strptime(entrada_inicio.get(), '%d/%m/%Y').date()
            data2 = datetime.strptime(entrada_fim.get(), '%d/%m/%Y').date()
            if data1 > data2:
                messagebox.showerror("Erro", "Data de início deve ser anterior à data final.")
                return
            root.datas = (data1, data2)
            root.destroy()
        except ValueError:
            messagebox.showerror("Erro", "Formato inválido. Use dd/mm/aaaa.")

    def on_cancelar():
        root.datas = None
        root.destroy()

    largura = 300
    altura = 120
    largura_tela = 1920
    altura_tela = 1080
    x = (largura_tela // 2) - (largura // 2)
    y = (altura_tela // 2) - (altura // 2)

    root = tk.Tk()
    root.title("Entrada de Datas")
    root.geometry(f"{largura}x{altura}+{x}+{y}")  # Centraliza a janela na tela
    root.resizable(False, False)
    root.attributes("-topmost", True)
    root.protocol("WM_DELETE_WINDOW", on_cancelar)  # Garante cancelamento limpo ao fechar

    tk.Label(root, text="Data de Início (dd/mm/aaaa):").grid(row=0, column=0, padx=10, pady=5)
    entrada_inicio = tk.Entry(root)
    entrada_inicio.grid(row=0, column=1, padx=10, pady=5)
    entrada_inicio.bind("<KeyRelease>", lambda e: formatar_data(e, entrada_inicio))

    tk.Label(root, text="Data Final (dd/mm/aaaa):").grid(row=1, column=0, padx=10, pady=5)
    entrada_fim = tk.Entry(root)
    entrada_fim.grid(row=1, column=1, padx=10, pady=5)
    entrada_fim.bind("<KeyRelease>", lambda e: formatar_data(e, entrada_fim))

    botao = tk.Button(root, text="Confirmar", command=on_confirmar)
    botao.grid(row=2, columnspan=2, pady=10)

    root.datas = None
    root.mainloop()

    return root.datas

def selecionar_arquivo_xlsx():
    root = tk.Tk() # Cria uma janela oculta
    root.withdraw()  # Oculta a janela principal
    global caminho_arquivo # Variável global para armazenar o caminho do arquivo
    caminho_arquivo = filedialog.askopenfilename( 
        title="Selecione um arquivo",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    ) # Abre o diálogo de seleção de arquivo
    root.destroy()  # Fecha a janela após a seleção

    return caminho_arquivo # Retorna o caminho do arquivo selecionado

def solicitar_entrada(titulo="Entrada de Dados", mensagem="Digite algo:"): 
    root = tk.Tk()  # Cria uma janela oculta
    root.withdraw()  # Oculta a janela principal

    resposta = simpledialog.askstring(titulo, mensagem) # Abre o diálogo de entrada de texto

    return resposta # Retorna a resposta do usuário

def bring_or_open_window_fullscreen(window_title, program_path):
    
    windows = gw.getWindowsWithTitle(window_title) # Obtém a lista de janelas abertas com o título fornecido
    
    if windows: # Se a janela já estiver aberta
        window = windows[0] # Seleciona a primeira janela encontrada
        
        if not window.isMaximized: # Se a janela não estiver maximizada
            window.restore() # Restaura a janela se estiver minimizada
            window.activate() # Traz a janela para o foco
            window.maximize() # Maximiza a janela
        else:
            window.activate() # Se a janela já estiver maximizada, apenas ativa
    else: # Se a janela não estiver aberta
        process = subprocess.Popen(program_path, shell=True) # Abre o programa
        time.sleep(3)  # Tempo para garantir que a janela seja criada
        
        windows = gw.getWindowsWithTitle(window_title) # Obtém a lista de janelas abertas novamente
        if windows: # Se a janela foi criada
            window = windows[0] # Seleciona a primeira janela encontrada
            window.restore() # Restaura a janela se estiver minimizada
            window.maximize() # Maximiza a janela
            window.activate() # Traz a janela para o foco
        mostrar_mensagem("Atenção", "Aguarde e faça login.", erro=False)

def close_specific_tab(window_title):
    windows = gw.getWindowsWithTitle(window_title) # Obtém a lista de janelas abertas com o título fornecido
    
    if windows: # Se a janela estiver aberta
        window = windows[0] # Seleciona a primeira janela encontrada
        window.activate()  # Traz a aba para o foco
        time.sleep(0.5)  # Pequena espera para garantir que está ativa

        # Para navegadores, usa Ctrl + W para fechar apenas a aba
        if "Google Chrome" in window_title or "Microsoft Edge" in window_title or "Mozilla Firefox" in window_title:
            pyautogui.hotkey('ctrl', 'w')
        else:
            window.close()  # Para outros programas, fecha a janela

def executar_script(script_path): # Executa o script Python e altera a interface visualmente
    global label_status # Variável global para o label de status

    # Verifica se label_status ainda existe antes de alterar
    if 'label_status' in globals() and label_status.winfo_exists(): # Se o label já existe, atualiza o texto e a cor
        label_status.config(text="Em execução", fg="white", bg="red")  # Texto branco com fundo vermelho
        root.config(bg="red")  # Fundo da janela vermelho
        root.update() # Atualiza a janela para refletir as mudanças imediatamente

    def run(): # Executa o script em uma thread separada
        try: # Executa o script
            cmd = ["python3", script_path] if os.name != "nt" else ["python", script_path] # Verifica o sistema operacional
            subprocess.run(cmd, check=True) # Executa o script e espera terminar
        except Exception as e: # Se ocorrer um erro
            if 'label_status' in globals() and label_status.winfo_exists(): # Verifica se o label ainda existe
                label_status.config(text="Erro!", fg="white", bg="darkred")  # Indica erro
        finally:
            if 'label_status' in globals() and label_status.winfo_exists():
                root.config(bg="SystemButtonFace")  # Volta ao fundo padrão
                label_status.config(text="", bg=root.cget("bg"))  # Remove o texto

    threading.Thread(target=run, daemon=True).start() # Inicia a thread para executar o script

def mostrar_mensagem(titulo, mensagem, erro=False): # Exibe uma mensagem de erro ou sucesso
    top = tk.Tk()
    top.withdraw()  # Oculta a janela principal
    top.attributes("-topmost", True)  # Mantém a caixa de diálogo no topo
    if erro:
        messagebox.showerror(titulo, mensagem)  # Mensagem de erro
    else:
        messagebox.showinfo(titulo, mensagem)  # Mensagem informativa

def exibir_popup():
    global popup, label_mensagem, botao_ok

    largura = 300
    altura = 120
    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()
    x = (largura_tela // 2) - (largura // 2)
    y = (altura_tela // 2) - (altura // 2)


    popup = tk.Toplevel()
    popup.title("Status")
    popup.geometry(f"{largura}x{altura}+{x}+{y}")  # Centraliza a janela na tela
    popup.resizable(False, False)
    popup.attributes("-topmost", True)
    popup.grab_set()  # Foca apenas nessa janela

    label_mensagem = tk.Label(popup, text="Iniciando...", font=("Arial", 11))
    label_mensagem.pack(pady=20)

    botao_ok = ttk.Button(popup, text="Okay", command=popup.destroy)
    botao_ok.pack(pady=10)
    botao_ok.pack_forget()  # Esconde o botão inicialmente

def atualizar_mensagem(texto, mostrar_ok=False):


    if label_mensagem and label_mensagem.winfo_exists():
        label_mensagem.config(text=texto)
    if mostrar_ok and botao_ok:
        botao_ok.pack(pady=10)  # Mostra o botão OK

#FUNÇÕES SECRETARIA

#FUNÇÕES FINANCEIRO     

def criarSheet_inadimplencia(nome, arquivo): # Cria uma nova aba com o nome do aluno
    if nome not in arquivo.sheetnames: # Verifica se a aba já existe
        arquivo.create_sheet(nome) # Cria a aba com o nome do aluno
        aba = arquivo[nome] # Abre a aba criada
        aba['A1'] = 'Nome' # Nome do aluno
        aba['B1'] = 'E-mail' # E-mail do aluno
        aba['C1'] = 'Numero do contrato' # Número do contrato
        aba['D1'] = 'Parcela' # Número da parcela
        aba['E1'] = 'Vencimento' # Data de vencimento
        aba['F1'] = 'Valor da Parcela' # Valor da parcela

def transferirDados_inadimplencia(linha_origem, aba_destino, aba_origem): # Transfere os dados da aba "Sheet" para a aba do aluno
    linha_destino = aba_destino.max_row + 1 # Última linha da aba do aluno
    for coluna in range(1, 7): # Colunas de 1 a 6
        if   coluna == 1:
            aba_destino.cell(row=linha_destino, column=coluna).value = aba_origem['I' + str(linha_origem)].value # Nome do aluno
        elif coluna == 2:
            aba_destino.cell(row=linha_destino, column=coluna).value = aba_origem['L' + str(linha_origem)].value # E-mail do aluno
        elif coluna == 3:
            aba_destino.cell(row=linha_destino, column=coluna).value = aba_origem['M' + str(linha_origem)].value # Número do contrato
        elif coluna == 4:
            aba_destino.cell(row=linha_destino, column=coluna).value = aba_origem['D' + str(linha_origem)].value # Número da parcela
        elif coluna == 5:
            aba_destino.cell(row=linha_destino, column=coluna).value = aba_origem['O' + str(linha_origem)].value # Data de vencimento
        elif coluna == 6:
            aba_destino.cell(row=linha_destino, column=coluna).value = aba_origem['W' + str(linha_origem)].value # Valor da parcela

def criar_arquivos_inadimplencia(caminho_arquivo):
    global nome_saida # Variável global para o nome do arquivo de saída
    datas = obter_intervalo_datas()  # Retorna tupla (data_inicio, data_fim)
    if not datas:
        print("Processo cancelado pelo usuário.")
        return

    data_inicio, data_fim = datas
    print(f"Intervalo de datas selecionado: {data_inicio} a {data_fim}")

    arquivo = load_workbook(caminho_arquivo)
    aba_basedados = arquivo['Sheet']
    ultima_linha = len(aba_basedados['A'])

    for linha in range(2, ultima_linha + 1):
        nome = aba_basedados['I' + str(linha)].value
        if not nome:
            continue
        nome = str(nome)[:31]

        data_vencimento = aba_basedados['O' + str(linha)].value
        if not data_vencimento:
            continue

        # Convertendo data de vencimento para datetime.date
        if isinstance(data_vencimento, str):
            try:
                data_vencimento = datetime.strptime(data_vencimento, "%d/%m/%Y")
            except ValueError:
                try:
                    data_vencimento = datetime.strptime(data_vencimento, "%Y-%m-%d")
                except ValueError:
                    continue

        if isinstance(data_vencimento, datetime):
            data_vencimento = data_vencimento.date()
        else:
            continue

        # Verifica se a data está dentro do intervalo
        if data_inicio <= data_vencimento <= data_fim:
            criarSheet_inadimplencia(nome, arquivo)
            aba_destino = arquivo[nome]
            transferirDados_inadimplencia(linha, aba_destino, aba_basedados)

    inicio_str = data_inicio.strftime("%d-%m-%Y") # Formata as datas para o nome do arquivo
    fim_str = data_fim.strftime("%d-%m-%Y") # Formata as datas para o nome do arquivo
    nome_arquivo_base = f"Inadimplentes_{inicio_str}_a_{fim_str}.xlsx" # Nome do arquivo com intervalo de datas
    nome_saida = salvar_com_nome_disponivel(nome_arquivo_base) # Salva o arquivo com um nome único
    arquivo.save(nome_saida)
    arquivo.close()
    print(f"Arquivo salvo como: {nome_saida}")

def criarTabela_emailInadimplencia(sheet, nome_arquivo): # Cria a tabela de e-mail com os dados da aba do aluno
    wb = load_workbook(nome_arquivo) # Abre o arquivo com as novas abas
    ws = wb[sheet] # Abre a aba do aluno
    linhas_tabela = "" # Inicializa a variável para armazenar as linhas da tabela

    for linha in range(2, ws.max_row + 1): # Percorre as linhas da aba do aluno
        num_contrato = ws.cell(row=linha, column=3).value # Número do contrato
        parcela = ws.cell(row=linha, column=4).value # Número da parcela
        vencimento = ws.cell(row=linha, column=5).value # Data de vencimento
        valor = ws.cell(row=linha, column=6).value # Valor da parcela
        linhas_tabela += f"<tr><td>{num_contrato}</td><td>{parcela}</td><td>{vencimento}</td><td>R$ {valor}</td></tr>\n" # Adiciona a linha à tabela

    return linhas_tabela # Retorna as linhas da tabela formatadas em HTML

def enviar_emails_inadimplecia(): # Envia os e-mails para os alunos
    global nome_saida # Variável global para o nome do arquivo de saída
    print("Iniciando o envio de e-mails...") # Mensagem de início
    print(f"planilha referência:{nome_saida}") # Mensagem de carregamento
    mostrar_mensagem("Atenção", "Aguarde, os e-mails estão sendo enviados.", erro=False) # Exibe mensagem de carregamento
    arquivo = load_workbook(nome_saida) # Abre o arquivo com as novas abas
    enviados = 0 # Contador de e-mails enviados
    naoenviados = 0 # Contador de e-mails não enviados

    with open("log.txt", "w", encoding="utf-8") as log: # Cria o arquivo de log
        log.write("Alunos sem e-mail registrado (e-mail não enviado):\n") # Cabeçalho do log
        
        for sheet in arquivo.sheetnames[1:]:  # Ignora a primeira aba (Sheet)
            aba = arquivo[sheet] # Abre a aba do aluno
            tabelaemail = criarTabela_emailInadimplencia(sheet, nome_saida) # Cria a tabela de e-mail com os dados da aba do aluno

            if not aba['B2'].value: # Se o e-mail estiver vazio
                naoenviados += 1 # Incrementa o contador de e-mails não enviados
                log.write(f"- {sheet}\n") # Adiciona o nome do aluno ao log
                print(f"E-mail não enviado para {aba['A2'].value} - {aba['B2'].value}") # Exibe mensagem de erro no console
            
            else: # Se o e-mail estiver preenchido

                
                outlook = win32com.client.Dispatch('outlook.application') # Cria o objeto Outlook
                mail = outlook.CreateItem(0) # Cria um novo e-mail
                mail.To = 'exemple@exemple.com.br'#aba['B2'].value  # Substitua para testes
                #mail.CC = 'exemple@exemple.com.br'  # Cópia para outro destinatário, se necessário
                mail.Subject = f"{aba['A2'].value} - Parcelas em aberto (TESTE)" # Assunto do e-mail
                #Corpo do e-mail em HTML
                mail.HTMLBody = f'''    
                <html>
                <body>
                    <p><strong>{aba['A2'].value}</strong></p>
                    <p>
                    Prezada(o) Cliente, Vimos por meio desta informar que até a presente data,
                    não acusamos em nossos registros o pagamento da(s) parcela(s) discriminada(s)
                    abaixo:
                    </p>
                    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
                    <thead>
                        <tr>
                        <th>Nº Contrato</th>
                        <th>Parcela</th>
                        <th>Vencimento</th>
                        <th>Valor da Parcela</th>
                        </tr>
                    </thead>
                    <tbody>
                        {tabelaemail}
                    </tbody>
                    </table>
                    <p>
                    Acreditando no sucesso da parceria entre V. Sª e esta Entidade, solicitamos
                    entrar em contato, pessoalmente no prazo de 10 dias ou através do Email:
                    exemple@exemple.com.br, para que possamos normalizar as pendências de
                    débitos constantes no seu contrato.
                    </p>
                    <p>Atenciosamente,</p>

                    <!-- Assinatura HTML -->
                    <div style="border-left: 4px solid #004b8d; padding-left: 15px; max-width: 500px; font-family: Arial, sans-serif;">
                    <div style="color: #004b8d; font-weight: bold; font-size: 16px;">Guilherme A. Villas Boas</div>
                    <div style="font-size: 14px;">Jovem Aprendiz Firjan SESI</div>
                    <div style="font-size: 14px;">(+55 21) 98151-8402</div>

                    <!-- Two side-by-side clickable images -->
                    <div style="margin-top: 10px; display: flex; gap: 10px;">
                        <a href="https://www.firjan.com.br" target="_blank">
                        <img src="https://iili.io/3UmPaXs.jpg" alt="Firjan">
                        </a>
                        <a href="https://www.firjan.com.br" target="_blank">
                        <img src="https://iili.io/3UmPcLG.jpg" alt="Great Place to Work 2022">
                        </a>
                    </div>

                    <div style="margin-top: 10px; font-size: 14px;">
                        <a href="https://www.firjan.com.br" target="_blank" style="color: #004b8d; text-decoration: none; font-weight: bold;">www.firjan.com.br</a>
                    </div>
                    </div>
                </body>
                </html>
                '''
                mail.Send() # Envia o e-mail
                enviados += 1 # Incrementa o contador de e-mails enviados
                print(f"E-mail enviado para {aba['A2'].value} - {aba['B2'].value}") # Exibe mensagem de sucesso no console
    arquivo.close() # Fecha o arquivo com as novas abas
    root = tk.Tk() # Cria uma janela oculta
    root.withdraw() # Oculta a janela principal
    root.attributes('-topmost', True) # Mantém a caixa de diálogo no topo
    messagebox.showinfo("Concluído", f"E-mails enviados com sucesso!\n\nTotal enviados: {enviados}\nTotal não enviados: {naoenviados}") # Exibe mensagem de sucesso

def iniciar_criacao_em_thread(caminho_arquivo):
    def tarefa():
        atualizar_mensagem("Gerando arquivo...")
        criar_arquivos_inadimplencia(caminho_arquivo)
        atualizar_mensagem("Arquivo gerado com sucesso!", mostrar_ok=True)

    exibir_popup()
    threading.Thread(target=tarefa).start()

#FUNÇÕES ATENDIMENTO
def criarSheet_atendimento(nome, arquivo): # Cria uma nova aba com o nome do aluno
    if nome not in arquivo.sheetnames: # Verifica se a aba já existe
        arquivo.create_sheet(nome) # Cria a aba com o nome do aluno
        aba = arquivo[nome] # Abre a aba criada
        aba['A1'] = 'Nome' # Nome do aluno
        aba['B1'] = 'E-mail' # E-mail do aluno
        aba['C1'] = 'Numero do contrato' # Número do contrato
        aba['D1'] = 'Parcela' # Número da parcela
        aba['E1'] = 'Vencimento' # Data de vencimento
        aba['F1'] = 'Valor da Parcela' # Valor da parcela

def criar_arquivos_processoAtendimento(caminho_arquivo): # Cria as abas e transfere os dados

    arquivo = load_workbook(caminho_arquivo) # Abre o arquivo original
    aba_basedados = arquivo.worksheets[0] # Abre a primeira aba
    ultima_linha = len(aba_basedados['A']) # Última linha da primeira aba

    for linha in range(2, ultima_linha + 1): #Cria as abas e transfere os dados
        nome = aba_basedados['D' + str(linha)].value # Nome do aluno
        criarSheet_atendimento(nome, arquivo) # Cria a aba com o nome do aluno
        aba_destino = arquivo[nome] # Abre a aba criada
        transferirDados_inadimplencia(linha, aba_destino, aba_basedados) # Transfere os dados da aba "Sheet" para a aba do aluno

    arquivo.save("2.xlsx") # Salva o arquivo com as novas abas
    arquivo.close() # Fecha o arquivo original

def listar_sheets(caminho_arquivo):
    try:
        wb = load_workbook(caminho_arquivo, read_only=True)
        print("Sheets encontradas:")
        for nome in wb.sheetnames:
            print(f"- {nome}")
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho_arquivo}")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def buscar_sheets_por_nome_parcial(caminho_arquivo):
    try:
        wb = load_workbook(caminho_arquivo)
        termo = input("Digite parte do nome da aba que deseja buscar: ").lower()

        print("\nSheets disponíveis:")
        for nome in wb.sheetnames:
            print(f"- {nome}")

        correspondencias = [wb[nome] for nome in wb.sheetnames if termo in nome.lower()]

        if correspondencias:
            print("\n✅ Abas encontradas que correspondem à busca:")
            for ws in correspondencias:
                print(f"- {ws.title}")
            return correspondencias
        else:
            print("\n❌ Nenhuma aba corresponde à busca.")
            return []

    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho_arquivo}")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

#TELAS PRINCIPAL
def showScreen_main(): 
    """Exibe a tela principal."""
    close_specific_tab("TOTVS Linha RM - Serviços  Alias: CorporeRM | 3-SENAI")
    for widget in root.winfo_children():
        widget.destroy()

    label_status = tk.Label(root, text="", bg=root.cget("bg"))
    label_status.pack(side="bottom", fill="x")


    # FRAME do título
    frame_titulo = tk.Frame(root, bg='#034AA6')
    frame_titulo.pack(pady=(20, 10))

    label_logo = tk.Label(frame_titulo, text="≡ SENAI ≡", bg='#034AA6', fg='white',
                      font=("Arial Black", 24, 'bold italic'))
    label_logo.pack()

    # FRAME do subtítulo
    frame_info = tk.Frame(root, bg='#034AA6')
    frame_info.pack(pady=5)

    label_info = tk.Label(frame_info, text="Escolha um setor", wraplength=200,
                      justify="center", font=("Arial", 12), bg='#034AA6', fg='white')
    label_info.pack()

    # FRAME 1
    frame1 = tk.Frame(root, bg='#034AA6')
    frame1.pack(pady=(10, 5))

    button1 = tk.Button(frame1, text="Secretaria", command=showScreen_secretaria,
                    height=1, width=15)
    button1.pack(side="left", padx=5)

    button2 = tk.Button(frame1, text="Financeiro", command=showScreen_financeiro,
                    height=1, width=15)
    button2.pack(side="left", padx=5)

    # FRAME 2
    frame2 = tk.Frame(root, bg='#034AA6')
    frame2.pack(pady=5)

    button3 = tk.Button(frame2, text="Atendimento", command=showScreen_Atendimento,
                    height=1, width=15)
    button3.pack(side="left", padx=5)

    button4 = tk.Button(frame2, text="Em Breve", command=None,
                    height=1, width=15)
    button4.pack(side="left", padx=5)

#TELAS SECRETARIA
def showScreen_secretaria():
    """Exibe a segunda tela."""
    for widget in root.winfo_children():
        widget.destroy()

    label_info = tk.Label(root, text="Escolha uma opção abaixo para continuar:", wraplength=200, justify="center")
    label_info.pack(pady=10)
    label_info.config(font=("Arial", 12),bg='#034AA6', fg='white')

    # Botão Código de Autenticação
    button1 = tk.Button(root, text="SGE", command=showScreen_sge, height=2, width=25)
    button1.pack(pady=5)

    # Botão Tela 2
    button2 = tk.Button(root, text="Em breve...", height=2, width=25)
    button2.pack(pady=5)

    back_button = tk.Button(root, text="Voltar", command=showScreen_main, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

def showScreen_sge():
    close_specific_tab(janela_mec)
    close_specific_tab(janela_mecLogin)
    """Exibe a segunda tela."""
    for widget in root.winfo_children():
        widget.destroy()

    label_info = tk.Label(root, text="Escolha uma opção abaixo para continuar:", wraplength=200, justify="center")
    label_info.config(font=("Arial", 12),bg='#034AA6', fg='white')
    label_info.pack(pady=10)

    # Botão Código de Autenticação
    button1 = tk.Button(root, text="MEC - SISTEC", command=showScreen_mec, height=2, width=25)
    button1.pack(pady=5)

    # Botão Tela 2
    button2 = tk.Button(root, text="Em breve...", height=2, width=25)
    button2.pack(pady=5)

    back_button = tk.Button(root, text="Voltar", command=showScreen_secretaria, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

    bring_or_open_window_fullscreen(janela_sge, "C:\Totvs\RM.NET\RM.exe")

def showScreen_mec():
    webbrowser.open(url_mec_login)

    for widget in root.winfo_children():
        widget.destroy()

    label = tk.Label(root, text="MEC - SISTEC", wraplength=200, justify="center")
    label.config(font=("Arial", 12),bg='#034AA6', fg='white')
    label.pack(pady=10)

    script_button1 = tk.Button(root, text="Código de Autenticação", command=lambda: executar_script("codigo-de-autenticacao-pyautogui.py"), height=2, width=20)
    script_button1.pack(pady=10)

    back_button = tk.Button(root, text="Voltar", command=showScreen_sge, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

    mostrar_mensagem("Atenção", "Faça login.", erro=False)

#TELAS FINANCEIRO
def showScreen_financeiro():
    """Exibe a segunda tela."""
    for widget in root.winfo_children():
        widget.destroy()

    label_info = tk.Label(root, text="Escolha uma opção abaixo para continuar:", wraplength=200, justify="center")
    label_info.config(font=("Arial", 12),bg='#034AA6', fg='white')
    label_info.pack(pady=10)
    
    # Botão 1
    button1 = tk.Button(root, text="Inadimplência", command=showScreen_inadimplencia, height=2, width=25)
    button1.pack(pady=10)

    # Botão 2
    button2 = tk.Button(root, text="Em breve", height=2, width=25)
    button2.pack(pady=10)

    back_button = tk.Button(root, text="Voltar", command=showScreen_main, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

def showScreen_inadimplencia():
    caminho_arquivo = selecionar_arquivo_xlsx()
    if caminho_arquivo:
        nome_arquivo = os.path.basename(caminho_arquivo)
        if not nome_arquivo.lower().endswith(".xlsx"):
            mostrar_mensagem("Erro", "Selecione um arquivo .xlsx válido.", erro=True)
            return
        nome_arquivo = os.path.splitext(nome_arquivo)[0] + ".xlsx"  # Adiciona a extensão .xlsx de volta
    else:
        mostrar_mensagem("Erro", "Nenhum arquivo selecionado.", erro=True)
        return
    
    for widget in root.winfo_children():
        widget.destroy()

    label_info = tk.Label(root, text="Inadimplência", wraplength=200, justify="center")
    label_info.config(font=("Arial Black", 12),bg='#034AA6', fg='white')
    label_info.pack(pady=10)

    label_arquivo = tk.Label(root, text=f"Planilha referência: {nome_arquivo}", wraplength=200, justify="center")
    label_arquivo.config(font=("Arial", 10),bg='#034AA6', fg='white')
    label_arquivo.pack(pady=1)

    # Botão 1
    button1 = tk.Button(root, text="Criar arquivos", command=lambda: iniciar_criacao_em_thread(caminho_arquivo), height=2, width=25)
    button1.pack(pady=10)

    # Botão 2
    button2 = tk.Button(root, text="Enviar e-mails", command=enviar_emails_inadimplecia, height=2, width=25)
    button2.pack(pady=10)

    back_button = tk.Button(root, text="Voltar", command=showScreen_financeiro, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

#TELAS ATENDIMENTO
def showScreen_Atendimento():
    """Exibe a tela de Atendimento."""
    for widget in root.winfo_children():
        widget.destroy()

    label_info = tk.Label(root, text="Escolha uma opção abaixo para continuar:", wraplength=200, justify="center")
    label_info.config(font=("Arial", 12),bg='#034AA6', fg='white')
    label_info.pack(pady=10)

    # Botão 1
    button1 = tk.Button(root, text="Processo Atendimento", command=showScreen_processoAtendimento, height=2, width=25)
    button1.pack(pady=10)

    # Botão 2
    button2 = tk.Button(root, text="Em breve...", height=2, width=25)
    button2.pack(pady=10)

    back_button = tk.Button(root, text="Voltar", command=showScreen_main, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

def showScreen_processoAtendimento():
    caminho_arquivo = selecionar_arquivo_xlsx()
    if caminho_arquivo:
        nome_arquivo = os.path.basename(caminho_arquivo)
        if not nome_arquivo.lower().endswith(".xlsx"):
            mostrar_mensagem("Erro", "Selecione um arquivo .xlsx válido.", erro=True)
            return
        nome_arquivo = os.path.splitext(nome_arquivo)[0] + ".xlsx"  # Adiciona a extensão .xlsx de volta
    else:
        mostrar_mensagem("Erro", "Nenhum arquivo selecionado.", erro=True)
        return
    
    for widget in root.winfo_children():
        widget.destroy()

    label_info = tk.Label(root, text="Atendimento", wraplength=200, justify="center")
    label_info.config(font=("Arial Black", 12),bg='#034AA6', fg='white')
    label_info.pack(pady=10)

    label_arquivo = tk.Label(root, text=f"Planilha referência: {nome_arquivo}", wraplength=200, justify="center")
    label_arquivo.config(font=("Arial", 10),bg='#034AA6', fg='white')
    label_arquivo.pack(pady=1)

    # Botão 1
    button1 = tk.Button(root, text="Listar", command=lambda: listar_sheets(caminho_arquivo), height=2, width=25)
    button1.pack(pady=10)

    # Botão 2
    button2 = tk.Button(root, text="Procurar", command=lambda: buscar_sheets_por_nome_parcial(caminho_arquivo), height=2, width=25)
    button2.pack(pady=10)

    back_button = tk.Button(root, text="Voltar", command=showScreen_Atendimento, height=1, width=10)
    back_button.pack(pady=10)
    back_button.place(relx=0.5, rely=0.9, anchor="center")

#MAIN

global label_status, nome_arquivo, caminho_arquivo # Variáveis globais para o label de status e o nome do arquivo
label_status = None # Inicializa a variável como None
nome_arquivo = None # Inicializa a variável como None
caminho_arquivo = None # Inicializa a variável como None  
# Criação da janela principal
root = tk.Tk()
root.title("SENApp")  # Título da janela
root.iconbitmap("img/icone.ico")  # Ícone da janela
root.geometry("300x250+1600+600")  # Posição inicial da janela
root.resizable(False, False)  # Janela não redimensionável
root.attributes("-topmost", True)  # Janela sempre no topo
root.configure(bg='#034AA6')  # Cor de fundo da janela

# Fecha o programa ao fechar a janela principal
root.protocol("WM_DELETE_WINDOW", root.quit)

# Exibe a tela inicial
showScreen_main()

# Inicia o loop da interface gráfica
root.mainloop()

close_specific_tab("[MEC - SisTec] - Google Chrome")
close_specific_tab("[MEC - SISTEC - v.4467 ] - Google Chrome")
close_specific_tab("TOTVS Linha RM - Serviços  Alias: CorporeRM | 3-SENAI")
